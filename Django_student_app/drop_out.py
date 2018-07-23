import django,os,sys
os.environ['DJANGO_SETTINGS_MODULE']="onlineproject.settings"
django.setup()
from onlineapp.models import *

import MySQLdb
import click
import openpyxl
from bs4 import BeautifulSoup


conn=MySQLdb.connect("localhost","deepak","deepak17")
cursor=conn.cursor()


del_sheet = openpyxl.load_workbook("students.xlsx")["Deletions"]
cursor.execute("use tutdb")
for row in range(2, del_sheet.max_row + 1):
    print(del_sheet.cell(row,4).value)
    c=Student(name=del_sheet.cell(row,1).value, college=College.objects.get(acronym=del_sheet.cell(row,2).value) ,email=del_sheet.cell(row,3).value, db_folder=del_sheet.cell(row,4).value)
    c.dropped_out=True
    c.save()

